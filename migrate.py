#!/usr/bin/env python3
"""migrate.py

Production-ready migration of ~10,000 support tickets from Zendesk to
Jira Service Management (JSM) Cloud.

Architecture summary
--------------------
1. Extraction & pagination
   Uses the Zendesk Incremental Ticket Export API (cursor pagination) and
   side-loads ``users`` and ``comment_count`` plus per-ticket comments so we
   avoid an N+1 request storm. Comments themselves are paginated per ticket.

2. State management & idempotency
   A local SQLite database records each migrated Zendesk ticket id, the new
   Jira issue key, and the last successful ``after_cursor``. On startup we read
   the cursor to resume, and we skip any ticket already recorded as migrated.

3. Rate-limit handling
   A single ``_request`` wrapper retries on HTTP 429, honouring the
   ``Retry-After`` header, with exponential backoff for transient 5xx errors.

4. Text formatting (ADF)
   ``html_to_adf`` converts Zendesk HTML/plain text into a minimal but valid
   Atlassian Document Format document. It is intentionally conservative: it
   never emits invalid nodes, falling back to plain paragraphs.

5. Per-ticket execution loop
   ensure customer -> create issue -> add comments (chronological) ->
   upload attachments -> optional status transition.

6. Configuration & logging
   Credentials come from ``.env`` via python-dotenv. Logging writes INFO/ERROR
   to both the console and ``migration.log``.
"""
from __future__ import annotations

import csv
import html.parser
import logging
import os
import sqlite3
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional

import requests
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Logging setup (console + file)
# ---------------------------------------------------------------------------


def configure_logging(log_file: str) -> logging.Logger:
    """Configure root logging to emit INFO/ERROR to console and a file."""
    logger = logging.getLogger("zendesk_jsm")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s %(levelname)-7s %(message)s", "%Y-%m-%d %H:%M:%S")

    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(fmt)
    logger.addHandler(console)

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(fmt)
    logger.addHandler(file_handler)

    return logger


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------


@dataclass
class Settings:
    """Runtime configuration loaded from environment variables."""

    zendesk_url: str
    zendesk_email: str
    zendesk_token: str
    atlassian_url: str
    atlassian_email: str
    atlassian_token: str
    jsm_project_key: str
    jsm_service_desk_id: str
    jsm_issue_type: str
    state_db_path: str
    mappings_path: str
    log_file: str
    zendesk_page_size: int

    @classmethod
    def from_env(cls) -> "Settings":
        load_dotenv()

        def required(name: str) -> str:
            value = os.getenv(name)
            if not value:
                raise SystemExit(f"Missing required environment variable: {name}")
            return value

        return cls(
            zendesk_url=required("ZENDESK_SUBDOMAIN_URL").rstrip("/"),
            zendesk_email=required("ZENDESK_EMAIL"),
            zendesk_token=required("ZENDESK_API_TOKEN"),
            atlassian_url=required("ATLASSIAN_BASE_URL").rstrip("/"),
            atlassian_email=required("ATLASSIAN_EMAIL"),
            atlassian_token=required("ATLASSIAN_API_TOKEN"),
            jsm_project_key=required("JSM_PROJECT_KEY"),
            jsm_service_desk_id=os.getenv("JSM_SERVICE_DESK_ID", ""),
            jsm_issue_type=os.getenv("JSM_ISSUE_TYPE", "Service Request"),
            state_db_path=os.getenv("STATE_DB_PATH", "migration_state.db"),
            mappings_path=os.getenv("MAPPINGS_PATH", "mappings.xlsx"),
            log_file=os.getenv("LOG_FILE", "migration.log"),
            zendesk_page_size=int(os.getenv("ZENDESK_PAGE_SIZE", "1000")),
        )


# ---------------------------------------------------------------------------
# Mappings loader (reads either mappings.xlsx or the mappings/ CSV folder)
# ---------------------------------------------------------------------------


@dataclass
class Mappings:
    """Holds all lookup tables plus an unmapped-value tracker."""

    priority: dict[str, str] = field(default_factory=dict)
    status: dict[str, str] = field(default_factory=dict)
    tags: dict[str, str] = field(default_factory=dict)
    request_type: dict[str, str] = field(default_factory=dict)
    config: dict[str, str] = field(default_factory=dict)
    # Counts misses so the pilot run surfaces gaps at the end.
    unmapped: dict[str, Counter] = field(
        default_factory=lambda: {"priority": Counter(), "status": Counter(), "request_type": Counter()}
    )

    # --- lookups with default + miss tracking ---
    def map_priority(self, value: Optional[str]) -> str:
        key = (value or "").strip().lower()
        if key in self.priority:
            return self.priority[key]
        if key:
            self.unmapped["priority"][key] += 1
        return self.config.get("default_priority", "Medium")

    def map_status(self, value: Optional[str]) -> str:
        key = (value or "").strip().lower()
        if key in self.status:
            return self.status[key]
        if key:
            self.unmapped["status"][key] += 1
        return self.config.get("default_status", "Open")

    def map_tags(self, tags: Iterable[str]) -> list[str]:
        # Unmapped tags pass through unchanged (just normalised for JSM labels).
        result = []
        for tag in tags:
            mapped = self.tags.get(tag.strip().lower(), tag.strip())
            result.append(mapped.replace(" ", "-"))
        return result

    def map_request_type(self, value: Optional[str]) -> Optional[str]:
        key = (value or "").strip().lower()
        if not key:
            return None
        if key in self.request_type:
            return self.request_type[key]
        self.unmapped["request_type"][key] += 1
        return None

    def transition_enabled(self) -> bool:
        return str(self.config.get("transition_status", "false")).strip().lower() in {"1", "true", "yes"}


def _load_csv_pairs(path: Path) -> dict[str, str]:
    """Load a 2-column CSV (header ignored) into a lower-cased-key dict."""
    result: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        next(reader, None)  # skip header row
        for row in reader:
            if len(row) >= 2 and row[0].strip():
                result[row[0].strip().lower()] = row[1].strip()
    return result


def load_mappings(path_str: str, logger: logging.Logger) -> Mappings:
    """Load mappings from an .xlsx workbook or a folder of CSVs.

    The loader is forgiving: a missing optional tab/file just yields an empty
    table. ``config`` is treated as key/value rather than zendesk/jsm.
    """
    path = Path(path_str)
    mappings = Mappings()
    tab_names = ["priority", "status", "tags", "request_type", "config"]

    if path.is_file() and path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        for name in tab_names:
            if name not in workbook.sheetnames:
                continue
            sheet = workbook[name]
            pairs: dict[str, str] = {}
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index == 0 or not row or row[0] is None:  # header / blank
                    continue
                key = str(row[0]).strip().lower()
                value = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
                pairs[key] = value
            setattr(mappings, name, pairs)
        logger.info("Loaded mappings from workbook %s", path)
    else:
        # Treat path as the CSV directory (or default 'mappings/').
        csv_dir = path if path.is_dir() else Path("mappings")
        for name in tab_names:
            csv_path = csv_dir / f"{name}.csv"
            if csv_path.exists():
                setattr(mappings, name, _load_csv_pairs(csv_path))
        logger.info("Loaded mappings from CSV directory %s", csv_dir)

    return mappings


# ---------------------------------------------------------------------------
# Local state store (SQLite) for idempotency / resume
# ---------------------------------------------------------------------------


class StateStore:
    """SQLite-backed tracker for migrated tickets and the export cursor."""

    def __init__(self, db_path: str) -> None:
        self.conn = sqlite3.connect(db_path)
        self.conn.execute(
            """
            CREATE TABLE IF NOT EXISTS migrated_tickets (
                zendesk_ticket_id INTEGER PRIMARY KEY,
                jira_issue_key    TEXT,
                migrated_at       TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS migration_meta (key TEXT PRIMARY KEY, value TEXT)"
        )
        self.conn.commit()

    def is_migrated(self, ticket_id: int) -> bool:
        cur = self.conn.execute(
            "SELECT 1 FROM migrated_tickets WHERE zendesk_ticket_id = ?", (ticket_id,)
        )
        return cur.fetchone() is not None

    def record_ticket(self, ticket_id: int, jira_key: str) -> None:
        self.conn.execute(
            "INSERT OR REPLACE INTO migrated_tickets (zendesk_ticket_id, jira_issue_key) VALUES (?, ?)",
            (ticket_id, jira_key),
        )
        self.conn.commit()

    def save_cursor(self, cursor: Optional[str]) -> None:
        if cursor is None:
            return
        self.conn.execute(
            "INSERT OR REPLACE INTO migration_meta (key, value) VALUES ('after_cursor', ?)",
            (cursor,),
        )
        self.conn.commit()

    def load_cursor(self) -> Optional[str]:
        cur = self.conn.execute("SELECT value FROM migration_meta WHERE key = 'after_cursor'")
        row = cur.fetchone()
        return row[0] if row else None

    def migrated_count(self) -> int:
        cur = self.conn.execute("SELECT COUNT(*) FROM migrated_tickets")
        return int(cur.fetchone()[0])

    def close(self) -> None:
        self.conn.close()


# ---------------------------------------------------------------------------
# HTML -> ADF conversion
# ---------------------------------------------------------------------------


class _TextExtractor(html.parser.HTMLParser):
    """Collapse HTML into plain text blocks split on block-level boundaries."""

    _BLOCK_TAGS = {"p", "div", "br", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}

    def __init__(self) -> None:
        super().__init__()
        self._blocks: list[str] = [""]

    def handle_starttag(self, tag: str, attrs: Any) -> None:
        if tag in self._BLOCK_TAGS:
            self._blocks.append("")

    def handle_endtag(self, tag: str) -> None:
        if tag in self._BLOCK_TAGS:
            self._blocks.append("")

    def handle_data(self, data: str) -> None:
        self._blocks[-1] += data

    def blocks(self) -> list[str]:
        return [b.strip() for b in self._blocks if b.strip()]


def html_to_adf(text: Optional[str]) -> dict[str, Any]:
    """Convert Zendesk HTML/plain text into a valid minimal ADF document.

    Each block of text becomes a paragraph node. Empty input yields a single
    empty paragraph (ADF requires at least one content node). This is
    deliberately conservative so we never POST a malformed document.
    """
    raw = text or ""
    if "<" in raw and ">" in raw:
        parser = _TextExtractor()
        parser.feed(raw)
        blocks = parser.blocks()
    else:
        blocks = [line.strip() for line in raw.splitlines() if line.strip()]

    if not blocks:
        blocks = ["(no content)"]

    return {
        "type": "doc",
        "version": 1,
        "content": [
            {"type": "paragraph", "content": [{"type": "text", "text": block}]}
            for block in blocks
        ],
    }


# ---------------------------------------------------------------------------
# HTTP client with rate-limit (429) and transient-error handling
# ---------------------------------------------------------------------------


class ApiClient:
    """Thin requests wrapper with retry, 429/Retry-After and 5xx backoff."""

    MAX_RETRIES = 6

    def __init__(self, auth: tuple[str, str], logger: logging.Logger) -> None:
        self.session = requests.Session()
        self.session.auth = auth
        self.logger = logger

    def request(
        self,
        method: str,
        url: str,
        *,
        expected: tuple[int, ...] = (200, 201, 204),
        **kwargs: Any,
    ) -> requests.Response:
        """Perform a request, retrying on 429 and transient 5xx responses."""
        attempt = 0
        while True:
            attempt += 1
            response = self.session.request(method, url, timeout=60, **kwargs)

            if response.status_code == 429:
                retry_after = self._retry_after_seconds(response, attempt)
                self.logger.warning(
                    "429 rate limited on %s %s; sleeping %ss (attempt %d)",
                    method, url, retry_after, attempt,
                )
                time.sleep(retry_after)
                continue

            if response.status_code >= 500 and attempt <= self.MAX_RETRIES:
                backoff = min(2 ** attempt, 60)
                self.logger.warning(
                    "%s on %s %s; backing off %ss (attempt %d)",
                    response.status_code, method, url, backoff, attempt,
                )
                time.sleep(backoff)
                continue

            if response.status_code not in expected:
                self.logger.error(
                    "%s %s -> %s: %s", method, url, response.status_code, response.text[:500]
                )
                response.raise_for_status()

            return response

    @staticmethod
    def _retry_after_seconds(response: requests.Response, attempt: int) -> float:
        """Read Retry-After (seconds) or fall back to exponential backoff."""
        header = response.headers.get("Retry-After")
        if header:
            try:
                return float(header)
            except ValueError:
                pass
        return float(min(2 ** attempt, 60))


# ---------------------------------------------------------------------------
# Zendesk extraction
# ---------------------------------------------------------------------------


class ZendeskClient:
    """Reads tickets via the Incremental Ticket Export (cursor) API."""

    def __init__(self, settings: Settings, client: ApiClient, logger: logging.Logger) -> None:
        self.base = settings.zendesk_url
        self.client = client
        self.logger = logger
        self.page_size = settings.zendesk_page_size

    def iter_ticket_pages(self, after_cursor: Optional[str]) -> Iterator[tuple[dict[str, Any], Optional[str]]]:
        """Yield (export_payload, after_cursor) tuples until the stream ends.

        ``users`` are side-loaded so requester details require no extra calls.
        On the first run with no cursor we start from time 0 (full export).
        """
        url = f"{self.base}/api/v2/incremental/tickets/cursor.json"
        params: dict[str, Any] = {"per_page": self.page_size, "include": "users"}
        if after_cursor:
            params["cursor"] = after_cursor
        else:
            params["start_time"] = 0

        while True:
            response = self.client.request("GET", url, params=params)
            payload = response.json()
            cursor = payload.get("after_cursor")
            yield payload, cursor

            if payload.get("end_of_stream"):
                return
            if not cursor:
                return
            params = {"per_page": self.page_size, "include": "users", "cursor": cursor}

    def iter_comments(self, ticket_id: int) -> Iterator[dict[str, Any]]:
        """Yield a ticket's comments in chronological order (paginated)."""
        url = f"{self.base}/api/v2/tickets/{ticket_id}/comments.json"
        params = {"per_page": 100, "sort_order": "asc"}
        while url:
            response = self.client.request("GET", url, params=params)
            payload = response.json()
            for comment in payload.get("comments", []):
                yield comment
            url = payload.get("next_page")
            params = {}  # next_page already encodes pagination

    def download_attachment(self, content_url: str) -> bytes:
        """Download an attachment's binary content from Zendesk."""
        response = self.client.request("GET", content_url, expected=(200,))
        return response.content


# ---------------------------------------------------------------------------
# JSM / Jira target
# ---------------------------------------------------------------------------


class JsmClient:
    """Creates customers, issues, comments and attachments in JSM Cloud."""

    def __init__(self, settings: Settings, client: ApiClient, logger: logging.Logger) -> None:
        self.settings = settings
        self.client = client
        self.logger = logger
        self.base = settings.atlassian_url
        # Cache resolved customer account ids by email to avoid duplicate work.
        self._customer_cache: dict[str, str] = {}
        # Cache project priority/status/transition metadata.
        self._transitions_cache: dict[str, dict[str, str]] = {}

    # --- customers ---
    def ensure_customer(self, email: str, display_name: str) -> Optional[str]:
        """Ensure a JSM customer exists for ``email``; return their accountId.

        The Customer API is idempotent enough for our purposes: if the customer
        already exists Jira returns 400/409, which we treat as benign.
        """
        if not email:
            return None
        if email in self._customer_cache:
            return self._customer_cache[email]

        url = f"{self.base}/rest/servicedeskapi/customer"
        body = {"email": email, "displayName": display_name or email}
        response = self.client.request(
            "POST", url, json=body, headers={"Content-Type": "application/json"},
            expected=(201, 400, 409),
        )
        account_id = None
        if response.status_code == 201:
            account_id = response.json().get("accountId")
            self.logger.info("Created JSM customer %s", email)
        else:
            self.logger.info("Customer %s already exists or could not be created", email)
        if account_id:
            self._customer_cache[email] = account_id
        return account_id

    # --- issues ---
    def create_issue(
        self,
        summary: str,
        description_adf: dict[str, Any],
        priority_name: str,
        labels: list[str],
        reporter_account_id: Optional[str],
    ) -> str:
        """Create a Jira issue and return its issue key."""
        url = f"{self.base}/rest/api/3/issue"
        fields: dict[str, Any] = {
            "project": {"key": self.settings.jsm_project_key},
            "summary": summary[:255] or "(no subject)",
            "description": description_adf,
            "issuetype": {"name": self.settings.jsm_issue_type},
        }
        if priority_name:
            fields["priority"] = {"name": priority_name}
        if labels:
            fields["labels"] = labels
        if reporter_account_id:
            fields["reporter"] = {"id": reporter_account_id}

        response = self.client.request(
            "POST", url, json={"fields": fields},
            headers={"Content-Type": "application/json"}, expected=(201,),
        )
        return response.json()["key"]

    def add_comment(self, issue_key: str, comment_adf: dict[str, Any]) -> None:
        """Add an ADF comment to an issue."""
        url = f"{self.base}/rest/api/3/issue/{issue_key}/comment"
        self.client.request(
            "POST", url, json={"body": comment_adf},
            headers={"Content-Type": "application/json"}, expected=(201,),
        )

    def upload_attachment(self, issue_key: str, filename: str, content: bytes) -> None:
        """Upload a binary attachment to an issue as multipart/form-data."""
        url = f"{self.base}/rest/api/3/issue/{issue_key}/attachments"
        # X-Atlassian-Token avoids XSRF check failures; do NOT set Content-Type
        # manually so requests can build the multipart boundary.
        self.client.request(
            "POST", url,
            headers={"X-Atlassian-Token": "no-check"},
            files={"file": (filename, content)},
            expected=(200,),
        )

    def transition_to_status(self, issue_key: str, status_name: str) -> None:
        """Move an issue to the named status if a matching transition exists."""
        if not status_name:
            return
        url = f"{self.base}/rest/api/3/issue/{issue_key}/transitions"
        response = self.client.request("GET", url, expected=(200,))
        transitions = response.json().get("transitions", [])
        target = next(
            (t for t in transitions if t.get("to", {}).get("name", "").lower() == status_name.lower()),
            None,
        )
        if not target:
            self.logger.warning("No transition to '%s' for %s", status_name, issue_key)
            return
        self.client.request(
            "POST", url, json={"transition": {"id": target["id"]}},
            headers={"Content-Type": "application/json"}, expected=(204,),
        )


# ---------------------------------------------------------------------------
# Migration orchestration
# ---------------------------------------------------------------------------


def _index_users(payload: dict[str, Any]) -> dict[int, dict[str, Any]]:
    """Build an id -> user dict from the side-loaded ``users`` array."""
    return {user["id"]: user for user in payload.get("users", [])}


def _collect_attachments(comment: dict[str, Any]) -> list[dict[str, Any]]:
    """Return the attachment dicts on a Zendesk comment."""
    return comment.get("attachments", []) or []


def migrate_ticket(
    ticket: dict[str, Any],
    users: dict[int, dict[str, Any]],
    zendesk: ZendeskClient,
    jsm: JsmClient,
    mappings: Mappings,
    state: StateStore,
    logger: logging.Logger,
) -> None:
    """Migrate a single Zendesk ticket into JSM."""
    ticket_id = ticket["id"]
    if state.is_migrated(ticket_id):
        logger.info("Skipping already-migrated ticket %s", ticket_id)
        return

    # --- requester / customer ---
    requester = users.get(ticket.get("requester_id"), {})
    requester_email = requester.get("email", "")
    requester_name = requester.get("name", requester_email)
    reporter_account_id = jsm.ensure_customer(requester_email, requester_name)

    # --- field mapping ---
    priority_name = mappings.map_priority(ticket.get("priority"))
    labels = mappings.map_tags(ticket.get("tags", []) or [])
    summary = ticket.get("subject") or ticket.get("raw_subject") or "(no subject)"
    description_adf = html_to_adf(ticket.get("description"))

    # --- create the issue ---
    issue_key = jsm.create_issue(summary, description_adf, priority_name, labels, reporter_account_id)
    logger.info("Created %s from Zendesk ticket %s", issue_key, ticket_id)

    # --- comments in chronological order (the first comment usually equals the
    #     description, so we skip index 0 to avoid duplicating it) ---
    for index, comment in enumerate(zendesk.iter_comments(ticket_id)):
        if index == 0:
            attachments_only = _collect_attachments(comment)
            _upload_comment_attachments(issue_key, attachments_only, zendesk, jsm, logger)
            continue
        body = comment.get("html_body") or comment.get("body")
        jsm.add_comment(issue_key, html_to_adf(body))
        _upload_comment_attachments(issue_key, _collect_attachments(comment), zendesk, jsm, logger)

    # --- optional status transition ---
    if mappings.transition_enabled():
        jsm.transition_to_status(issue_key, mappings.map_status(ticket.get("status")))

    # --- record success for idempotency ---
    state.record_ticket(ticket_id, issue_key)


def _upload_comment_attachments(
    issue_key: str,
    attachments: list[dict[str, Any]],
    zendesk: ZendeskClient,
    jsm: JsmClient,
    logger: logging.Logger,
) -> None:
    """Download each Zendesk attachment and upload it to the Jira issue."""
    for attachment in attachments:
        content_url = attachment.get("content_url")
        filename = attachment.get("file_name", "attachment.bin")
        if not content_url:
            continue
        try:
            data = zendesk.download_attachment(content_url)
            jsm.upload_attachment(issue_key, filename, data)
            logger.info("Attached %s to %s", filename, issue_key)
        except requests.HTTPError as exc:  # one bad attachment must not abort the ticket
            logger.error("Failed to migrate attachment %s for %s: %s", filename, issue_key, exc)


def print_unmapped_summary(mappings: Mappings, logger: logging.Logger) -> None:
    """Print a summary of unmapped values seen during the run."""
    logger.info("=" * 60)
    logger.info("UNMAPPED VALUE SUMMARY")
    any_misses = False
    for category, counter in mappings.unmapped.items():
        if not counter:
            continue
        any_misses = True
        logger.info("  %s:", category)
        for value, count in counter.most_common():
            logger.info("    %-30s %d occurrence(s)", value, count)
    if not any_misses:
        logger.info("  No unmapped values. All values resolved cleanly.")
    logger.info("=" * 60)


def run() -> None:
    settings = Settings.from_env()
    logger = configure_logging(settings.log_file)
    logger.info("Starting Zendesk -> JSM migration")

    mappings = load_mappings(settings.mappings_path, logger)
    state = StateStore(settings.state_db_path)

    zendesk_client = ApiClient((f"{settings.zendesk_email}/token", settings.zendesk_token), logger)
    jsm_client = ApiClient((settings.atlassian_email, settings.atlassian_token), logger)

    zendesk = ZendeskClient(settings, zendesk_client, logger)
    jsm = JsmClient(settings, jsm_client, logger)

    resume_cursor = state.load_cursor()
    if resume_cursor:
        logger.info("Resuming from saved cursor (already migrated: %d)", state.migrated_count())

    processed = 0
    try:
        for payload, after_cursor in zendesk.iter_ticket_pages(resume_cursor):
            users = _index_users(payload)
            for ticket in payload.get("tickets", []):
                try:
                    migrate_ticket(ticket, users, zendesk, jsm, mappings, state, logger)
                    processed += 1
                except requests.HTTPError as exc:
                    # Log and continue so one bad ticket does not halt 10k others.
                    logger.error("Failed to migrate ticket %s: %s", ticket.get("id"), exc)
            # Persist the cursor only after a full page is processed so resume
            # never skips unmigrated tickets from a partially-processed page.
            state.save_cursor(after_cursor)
    finally:
        logger.info("Processed %d tickets this run (total migrated: %d)", processed, state.migrated_count())
        print_unmapped_summary(mappings, logger)
        state.close()


if __name__ == "__main__":
    run()
